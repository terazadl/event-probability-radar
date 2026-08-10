#!/usr/bin/env python3
"""资金流层快照（美股 + BTC）

和 fetch_fred_snapshot.py 是同一套工具的两块：
- fetch_fred_snapshot.py 回答"利率和信用环境怎么样"（价格/贴现率层）
- 本脚本回答"钱在往哪里搬"（配置层）

设计前提：基本面告诉你约束在哪，资金流才告诉你价格往哪走。
两层都要看，任何一层单独用都会系统性犯错。

用法：
    .venv/bin/python3 Scripts/fetch_flows_snapshot.py

可选依赖（只有 FINRA 融资余额需要）：
    pip3 install openpyxl
"""
from __future__ import annotations

import csv
import io
import json
import re
import subprocess
import sys
from bisect import bisect_right
from dataclasses import dataclass
from datetime import date, datetime, timedelta
from pathlib import Path


SYSTEM_DIR = Path(__file__).resolve().parents[1]
DATA_DIR = SYSTEM_DIR / "Data" / "flows"
REPORT_DIR = SYSTEM_DIR / "Reports"

CURL_TIMEOUT = "60"
CURL_RETRIES = "3"


# ---------------------------------------------------------------- 数据结构

@dataclass(frozen=True)
class Metric:
    key: str
    name: str
    category: str
    unit: str
    kind: str            # "level" = 存量，看变化；"flow" = 流量，看窗口累计
    windows: tuple       # 要报告的窗口（天）
    why: str
    source: str
    caveat: str = ""


METRICS = [
    Metric(
        key="NETLIQ",
        name="美联储净流动性",
        category="总量流动性",
        unit="trillions_usd",
        kind="level",
        windows=(28, 91),
        why="WALCL−TGA−RRP。风险资产最上游的水位，领先于估值扩张/收缩",
        source="FRED WALCL / WTREGEN / RRPONTSYD",
        caveat="周频，且是会计口径不是交易口径；只看方向和斜率，不要看绝对值",
    ),
    Metric(
        key="RRPONTSYD",
        name="隔夜逆回购余额",
        category="总量流动性",
        unit="billions_usd",
        kind="level",
        windows=(28, 91),
        why="已经基本抽干；再度回升说明系统里的钱找不到去处",
        source="FRED RRPONTSYD",
    ),
    Metric(
        key="WRMFNS",
        name="零售货币基金规模",
        category="场外现金",
        unit="billions_usd",
        kind="level",
        windows=(28, 365),
        why="场外干火药。规模仍在增而股指新高≠见顶；规模掉头下行才是钱真的进场",
        source="FRED WRMFNS（周频，非季调）",
        caveat="口径只含零售，机构货基自 2021 年起 FRED 已停更",
    ),
    Metric(
        key="MARGIN",
        name="FINRA 融资余额",
        category="杠杆",
        unit="billions_usd",
        kind="level",
        windows=(365,),
        why="散户+经纪商杠杆。同比增速比绝对值有用得多，是波动率的燃料",
        source="FINRA Margin Statistics（月频，滞后约一个月）",
        caveat="需要 openpyxl；FINRA 的文件路径偶尔变动，失败属正常",
    ),
    Metric(
        key="STABLECOIN",
        name="稳定币总市值",
        category="加密内部资金",
        unit="billions_usd",
        kind="level",
        windows=(30, 90),
        why="加密世界的 M2。净增发是 BTC 最干净的领先资金指标之一",
        source="DefiLlama /stablecoincharts/all",
        caveat="含新链新币上线带来的机械式增量，单月异常值要人工看一眼",
    ),
    Metric(
        key="BTCETF",
        name="BTC 现货 ETF 净流入",
        category="加密内部资金",
        unit="millions_usd",
        kind="flow",
        windows=(5, 20),
        why="传统资金进出加密的唯一干净通道，日频、无滞后",
        source="Farside Investors farside.co.uk/btc/",
        caveat="HTML 抓取，页面改版即失效；失败时手动看一眼即可",
    ),
]


# ---------------------------------------------------------------- 通用工具

def _curl_command(url: str, *, browser_user_agent: bool = False) -> list[str]:
    command = [
        "curl", "-L", "--http1.1", "--ipv4", "-fsSL",
        "--connect-timeout", "15", "--max-time", CURL_TIMEOUT,
        "--retry", CURL_RETRIES, "--retry-delay", "1", "--retry-all-errors",
    ]
    if browser_user_agent:
        command.extend(["-A", "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7)"])
    command.append(url)
    return command


def curl(url: str, *, browser_user_agent: bool = False) -> str:
    result = subprocess.run(
        _curl_command(url, browser_user_agent=browser_user_agent),
        check=True,
        capture_output=True,
        text=True,
    )
    return result.stdout


def curl_binary(url: str, *, browser_user_agent: bool = False) -> bytes:
    result = subprocess.run(
        _curl_command(url, browser_user_agent=browser_user_agent),
        check=True,
        capture_output=True,
    )
    return result.stdout


def parse_fred_csv(raw_csv: str) -> list[tuple[date, float]]:
    reader = csv.DictReader(raw_csv.splitlines())
    observations: list[tuple[date, float]] = []
    fieldnames = reader.fieldnames or []
    if len(fieldnames) < 2:
        return observations
    date_key, value_key = fieldnames[0], fieldnames[1]
    for row in reader:
        raw_value = (row.get(value_key) or "").strip()
        if raw_value in {"", "."}:
            continue
        try:
            value = float(raw_value)
            obs_date = datetime.strptime(row[date_key], "%Y-%m-%d").date()
        except (ValueError, TypeError, KeyError):
            continue
        observations.append((obs_date, value))
    observations.sort()
    return observations


def value_on_or_before(observations, target: date):
    dates = [item[0] for item in observations]
    idx = bisect_right(dates, target) - 1
    if idx < 0:
        return None
    return observations[idx]


def sum_window(observations, end: date, days: int) -> float:
    start = end - timedelta(days=days)
    return sum(v for d, v in observations if start < d <= end)


# ---------------------------------------------------------------- 各数据源

def fetch_fred(series_id: str) -> list[tuple[date, float]]:
    # The longest comparison window is one year. Limiting the response to two
    # years makes the daily RRP series much less likely to time out. FRED also
    # behaves more reliably without a forged browser User-Agent.
    start = date.today() - timedelta(days=730)
    url = (
        "https://fred.stlouisfed.org/graph/fredgraph.csv"
        f"?id={series_id}&cosd={start.isoformat()}&coed={date.today().isoformat()}"
    )
    raw = curl(url)
    (DATA_DIR / f"{series_id}.csv").write_text(raw, encoding="utf-8")
    return parse_fred_csv(raw)


def build_net_liquidity(walcl, tga, rrp) -> list[tuple[date, float]]:
    """WALCL(百万) − TGA(十亿)*1000 − RRP(十亿)*1000，返回单位：万亿美元"""
    out: list[tuple[date, float]] = []
    for d, v in walcl:
        t = value_on_or_before(tga, d)
        r = value_on_or_before(rrp, d)
        if t is None or r is None:
            continue
        millions = v - (t[1] + r[1]) * 1000.0
        out.append((d, millions / 1_000_000.0))
    return out


def fetch_stablecoin_mcap() -> list[tuple[date, float]]:
    """DefiLlama 全量稳定币市值，返回单位：十亿美元"""
    raw = curl("https://stablecoins.llama.fi/stablecoincharts/all")
    (DATA_DIR / "stablecoins_raw.json").write_text(raw, encoding="utf-8")
    payload = json.loads(raw)
    out: list[tuple[date, float]] = []
    for point in payload:
        try:
            ts = int(point["date"])
            circ = point.get("totalCirculatingUSD") or {}
            if isinstance(circ, dict):
                total = sum(float(x) for x in circ.values())
            else:
                total = float(circ)
        except (KeyError, TypeError, ValueError):
            continue
        out.append((datetime.utcfromtimestamp(ts).date(), total / 1e9))
    out.sort()
    return out


_NUM_RE = re.compile(r"^\(?-?[\d,]+(?:\.\d+)?\)?$")


def _to_float(text: str):
    text = text.strip().replace(",", "")
    if text in {"", "-", "—", "n/a"}:
        return 0.0
    negative = text.startswith("(") and text.endswith(")")
    text = text.strip("()")
    try:
        value = float(text)
    except ValueError:
        return None
    return -value if negative else value


def fetch_btc_etf_flows() -> list[tuple[date, float]]:
    """Farside 日频净流入，返回单位：百万美元。HTML 抓取，脆弱。"""
    html = curl("https://farside.co.uk/btc/", browser_user_agent=True)
    (DATA_DIR / "farside_btc_raw.html").write_text(html, encoding="utf-8")
    rows = re.findall(r"<tr[^>]*>(.*?)</tr>", html, re.S | re.I)
    out: list[tuple[date, float]] = []
    for row in rows:
        cells = [
            re.sub(r"<[^>]+>", "", c).replace("&nbsp;", " ").strip()
            for c in re.findall(r"<t[dh][^>]*>(.*?)</t[dh]>", row, re.S | re.I)
        ]
        if len(cells) < 3:
            continue
        try:
            obs_date = datetime.strptime(cells[0], "%d %b %Y").date()
        except ValueError:
            continue
        total = None
        for cell in reversed(cells[1:]):
            if _NUM_RE.match(cell.replace(" ", "")):
                total = _to_float(cell)
                break
        if total is None:
            continue
        out.append((obs_date, total))
    out.sort()
    return out


def fetch_finra_margin() -> list[tuple[date, float]]:
    """FINRA 融资余额（debit balances），返回单位：十亿美元。"""
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise RuntimeError("需要 openpyxl：pip3 install openpyxl") from exc

    candidates = [
        "https://www.finra.org/sites/default/files/margin-statistics.xlsx",
        "https://www.finra.org/sites/default/files/2021-03/margin-statistics.xlsx",
    ]
    blob = None
    last_error = None
    for url in candidates:
        try:
            blob = curl_binary(url, browser_user_agent=True)
            break
        except Exception as exc:  # noqa: BLE001
            last_error = exc
    if blob is None:
        raise RuntimeError(f"FINRA 文件下载失败（路径可能已变动）：{last_error}")

    (DATA_DIR / "finra_margin.xlsx").write_bytes(blob)
    workbook = load_workbook(io.BytesIO(blob), data_only=True)
    sheet = workbook[workbook.sheetnames[0]]

    out: list[tuple[date, float]] = []
    for row in sheet.iter_rows(values_only=True):
        if not row or row[0] is None:
            continue
        raw_date, raw_value = row[0], row[1] if len(row) > 1 else None
        obs_date = None
        if isinstance(raw_date, datetime):
            obs_date = raw_date.date()
        elif isinstance(raw_date, date):
            obs_date = raw_date
        elif isinstance(raw_date, str):
            for fmt in ("%b-%y", "%B %Y", "%Y-%m", "%b %Y"):
                try:
                    obs_date = datetime.strptime(raw_date.strip(), fmt).date()
                    break
                except ValueError:
                    continue
        if obs_date is None or not isinstance(raw_value, (int, float)):
            continue
        # FINRA 原表单位是百万美元
        out.append((obs_date, float(raw_value) / 1000.0))
    out.sort()
    if not out:
        raise RuntimeError("FINRA 表格结构无法解析，请手动核对")
    return out


# ---------------------------------------------------------------- 格式化

def fmt_value(value: float, unit: str) -> str:
    if unit == "trillions_usd":
        return f"${value:,.3f}T"
    if unit == "billions_usd":
        return f"${value:,.1f}B"
    if unit == "millions_usd":
        return f"${value:,.0f}M"
    return f"{value:,.2f}"


def fmt_change(delta, unit: str, pct=None) -> str:
    if delta is None:
        return "n/a"
    if unit == "trillions_usd":
        base = f"${delta:+,.3f}T"
    elif unit == "billions_usd":
        base = f"${delta:+,.1f}B"
    elif unit == "millions_usd":
        base = f"${delta:+,.0f}M"
    else:
        base = f"{delta:+,.2f}"
    if pct is not None:
        base += f"（{pct:+.1f}%）"
    return base


def window_label(days: int, kind: str) -> str:
    verb = "累计" if kind == "flow" else "变化"
    if days % 365 == 0:
        return f"{days // 365}年{verb}"
    if days >= 28 and days % 7 == 0:
        return f"{days // 7}周{verb}"
    return f"{days}日{verb}"


# ---------------------------------------------------------------- 解读层

def quick_read(results: dict) -> list[str]:
    """规则写在这里，方便以后一条条回测和淘汰。
    每条规则都应该能被证伪——如果一条规则连续多次给出反向信号，就删掉它。"""
    notes: list[str] = []
    ok = {k: v for k, v in results.items() if not v.get("error")}

    netliq = ok.get("NETLIQ", {}).get("changes", {}).get(28, {}).get("delta")
    stable = ok.get("STABLECOIN", {}).get("changes", {}).get(30, {}).get("delta")
    etf20 = ok.get("BTCETF", {}).get("changes", {}).get(20, {}).get("delta")
    mmf = ok.get("WRMFNS", {}).get("changes", {}).get(28, {}).get("delta")
    margin_pct = ok.get("MARGIN", {}).get("changes", {}).get(365, {}).get("pct")

    if netliq is not None and stable is not None:
        if netliq > 0 and stable > 0:
            notes.append("净流动性和稳定币同向扩张：风险资产处在顺风环境，回调更可能是噪音。")
        elif netliq < 0 and stable < 0:
            notes.append("净流动性和稳定币同向收缩：这是资金层最值得警惕的组合，风险资产的回调更可能有延续性。")
        else:
            notes.append("净流动性和稳定币方向背离：传统与加密两个池子的钱在分道走，不要用单一叙事解释行情。")

    if stable is not None and stable < 0:
        notes.append("稳定币净增发转负，加密内部新钱在流出，BTC 的上行通常需要这一项先转正。")

    if etf20 is not None:
        if etf20 > 0 and stable is not None and stable < 0:
            notes.append("ETF 在净流入但稳定币在收缩：买盘偏传统资金、加密原生资金缺席，这种结构历史上不太稳。")
        elif etf20 < 0:
            notes.append("BTC 现货 ETF 近 20 个交易日净流出，传统资金在减配。")

    if mmf is not None:
        if mmf > 0:
            notes.append("货币基金规模仍在增长：场外现金还在积累，指数新高本身不构成见顶理由。")
        else:
            notes.append("货币基金规模掉头向下：钱正在离开现金，注意这通常发生在行情中后段而不是起点。")

    if margin_pct is not None:
        if margin_pct > 40:
            notes.append(f"融资余额同比 {margin_pct:+.0f}%，杠杆偏热。这不是择时信号，但它意味着一旦回调，波动会被放大。")
        elif margin_pct < -10:
            notes.append(f"融资余额同比 {margin_pct:+.0f}%，杠杆在主动收缩，通常伴随一段情绪修复期。")

    if not notes:
        notes.append("本次没有触发任何规则，作为基准数据留存即可。")
    return notes


def build_alerts(results: dict) -> list[str]:
    """只有这些情况值得打断你——其余的进 markdown 就好。"""
    alerts: list[str] = []
    ok = {k: v for k, v in results.items() if not v.get("error")}

    netliq = ok.get("NETLIQ", {}).get("changes", {}).get(28, {}).get("delta")
    stable = ok.get("STABLECOIN", {}).get("changes", {}).get(30, {}).get("delta")
    stable_prev = ok.get("STABLECOIN", {}).get("changes", {}).get(90, {}).get("delta")
    etf5 = ok.get("BTCETF", {}).get("changes", {}).get(5, {}).get("delta")
    margin_pct = ok.get("MARGIN", {}).get("changes", {}).get(365, {}).get("pct")

    if netliq is not None and stable is not None and netliq < 0 and stable < 0:
        alerts.append("【双收缩】净流动性 4 周为负 且 稳定币 30 日净增发为负")
    if stable is not None and stable_prev is not None and stable < 0 < stable_prev:
        alerts.append("【拐点】稳定币净增发由正转负")
    if etf5 is not None and etf5 < -1500:
        alerts.append(f"【急流出】BTC 现货 ETF 近 5 日净流出 {abs(etf5):,.0f}M")
    if margin_pct is not None and margin_pct > 50:
        alerts.append(f"【杠杆过热】FINRA 融资余额同比 {margin_pct:+.0f}%")
    return alerts


# ---------------------------------------------------------------- 报告

def build_report(results: dict, alerts: list[str]) -> str:
    today = date.today().isoformat()
    success = {k: v for k, v in results.items() if not v.get("error")}
    failed = {k: v for k, v in results.items() if v.get("error")}

    lines = [
        f"# 资金流快照 - {today}",
        "",
        f"日期：{today}",
        "标签：#资金流 #流动性 #BTC #美股 #投研系统",
        "",
        "## 一句话",
        "",
        "这一层不回答'经济好不好'，只回答'钱在往哪搬'。它和 [[Reports/FRED 利率与通胀快照 " + today + "|利率快照]] 配合看：",
        "利率决定贴现率，资金流决定谁在出价。",
        "",
    ]

    if alerts:
        lines.extend(["## 触发提醒", ""])
        lines.extend(f"- {a}" for a in alerts)
        lines.append("")

    lines.extend(["## 快速解读", ""])
    lines.extend(f"- {n}" for n in quick_read(results))

    lines.extend(["", "## 指标表", "", "| 类别 | 指标 | 最新日期 | 最新值 | 变化 | 为什么看 |", "|---|---|---:|---:|---|---|"])
    for metric in METRICS:
        row = success.get(metric.key)
        if not row:
            continue
        change_parts = []
        for days in metric.windows:
            change = row["changes"].get(days, {})
            change_parts.append(
                f"{window_label(days, metric.kind)} {fmt_change(change.get('delta'), metric.unit, change.get('pct'))}"
            )
        lines.append(
            f"| {metric.category} | {metric.name} | {row['latest_date']} | "
            f"{fmt_value(row['latest_value'], metric.unit)} | {'<br>'.join(change_parts)} | {metric.why} |"
        )

    if failed:
        lines.extend(["", "## 抓取失败", ""])
        for key, row in failed.items():
            metric = next(m for m in METRICS if m.key == key)
            lines.append(f"- **{metric.name}**（{metric.source}）：{row['error']}")

    lines.extend(["", "## 口径与陷阱", ""])
    for metric in METRICS:
        if metric.caveat:
            lines.append(f"- **{metric.name}**：{metric.caveat}")

    lines.extend([
        "",
        "## 怎么用",
        "",
        "1. 资金流层**不产生买卖信号**，它只调整你对基本面判断的仓位权重。",
        "2. 基本面看空 + 资金流扩张 = 少数最容易亏钱的组合，此时不要重仓做空，也不要嘴硬。",
        "3. 基本面看空 + 资金流收缩 = 才是可以下手的时候。",
        "4. 连续两次快照指向同一方向，才写进 [[观点与假设库]]。",
        "5. 如果某条规则连续给出反向信号，去 `quick_read()` 里把它删掉——规则库要能减不能只增。",
        "",
        "## 数据来源",
        "",
    ])
    seen = set()
    for metric in METRICS:
        if metric.source not in seen:
            lines.append(f"- {metric.name}：{metric.source}")
            seen.add(metric.source)

    return "\n".join(lines) + "\n"


# ---------------------------------------------------------------- 主流程

def compute_row(observations, metric: Metric) -> dict:
    latest_date, latest_value = observations[-1]
    changes: dict[int, dict] = {}
    for days in metric.windows:
        if metric.kind == "flow":
            changes[days] = {"delta": sum_window(observations, latest_date, days), "pct": None}
        else:
            past = value_on_or_before(observations, latest_date - timedelta(days=days))
            if past is None:
                changes[days] = {"delta": None, "pct": None}
            else:
                delta = latest_value - past[1]
                pct = (delta / abs(past[1]) * 100.0) if past[1] else None
                changes[days] = {"delta": delta, "pct": pct}
    return {
        "key": metric.key,
        "name": metric.name,
        "latest_date": latest_date.isoformat(),
        "latest_value": latest_value,
        "changes": changes,
        "error": "",
    }


def main() -> None:
    DATA_DIR.mkdir(parents=True, exist_ok=True)
    REPORT_DIR.mkdir(parents=True, exist_ok=True)

    series: dict[str, list] = {}
    results: dict[str, dict] = {}

    def run(key: str, loader) -> None:
        metric = next(m for m in METRICS if m.key == key)
        try:
            observations = loader()
            if not observations:
                raise RuntimeError("没有解析到任何观测值")
            series[key] = observations
            results[key] = compute_row(observations, metric)
        except Exception as exc:  # noqa: BLE001
            results[key] = {"key": key, "name": metric.name, "error": str(exc)}

    # FRED 三件套先单独抓，净流动性要用
    fred_raw: dict[str, list] = {}
    fred_errors: dict[str, str] = {}
    for series_id in ("WALCL", "WTREGEN", "RRPONTSYD"):
        try:
            fred_raw[series_id] = fetch_fred(series_id)
        except Exception as exc:  # noqa: BLE001
            fred_errors[series_id] = str(exc)

    def load_netliq():
        if fred_errors:
            detail = "; ".join(
                f"{series_id}: {message}"
                for series_id, message in fred_errors.items()
            )
            raise RuntimeError(detail)
        return build_net_liquidity(
            fred_raw.get("WALCL") or [],
            fred_raw.get("WTREGEN") or [],
            fred_raw.get("RRPONTSYD") or [],
        )

    run("NETLIQ", load_netliq)

    def load_rrp():
        if "RRPONTSYD" in fred_errors:
            raise RuntimeError(f"RRPONTSYD: {fred_errors['RRPONTSYD']}")
        return fred_raw.get("RRPONTSYD") or []

    run("RRPONTSYD", load_rrp)
    run("WRMFNS", lambda: fetch_fred("WRMFNS"))
    run("STABLECOIN", fetch_stablecoin_mcap)
    run("BTCETF", fetch_btc_etf_flows)
    run("MARGIN", fetch_finra_margin)

    alerts = build_alerts(results)
    today = date.today().isoformat()

    serialisable = {
        k: {kk: vv for kk, vv in v.items() if kk != "changes"} | {
            "changes": {str(d): c for d, c in v.get("changes", {}).items()}
        }
        for k, v in results.items()
    }
    (DATA_DIR / "latest.json").write_text(
        json.dumps(serialisable, ensure_ascii=False, indent=2), encoding="utf-8"
    )
    (DATA_DIR / "alerts.json").write_text(
        json.dumps({"date": today, "alerts": alerts}, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )

    if not any(not row.get("error") for row in results.values()):
        raise SystemExit("所有数据源都失败了，检查网络或看 Data/flows/latest.json")

    report_path = REPORT_DIR / f"资金流快照 {today}.md"
    report_path.write_text(build_report(results, alerts), encoding="utf-8")
    print(report_path)
    for alert in alerts:
        print("ALERT:", alert)

    failed_results = [row for row in results.values() if row.get("error")]
    if failed_results:
        failed_names = "、".join(row["name"] for row in failed_results)
        print(
            f"部分资金流指标失败（{len(failed_results)}/{len(results)}）：{failed_names}",
            file=sys.stderr,
        )
        raise SystemExit(2)


if __name__ == "__main__":
    main()
